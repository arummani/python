#!/usr/bin/env python3
"""
Fetch newly added Indian movies, web-series, and documentaries from major
streaming platforms using the OTT Details API on RapidAPI.
Optionally email the results as an Excel attachment via Gmail SMTP.

Platforms queried:
  US region : Netflix, Prime Video, Hulu, Zee5
  IN region : Netflix, Prime Video, Hotstar, Zee5

Data flow:
  1. /getnew   — paginated new arrivals per region
  2. /getTitleDetails — enrich each title with IMDb rating & genres

Usage:
    export OTT_DETAILS_API_KEY="your-rapidapi-key"

    # Optional – enable email delivery:
    export SENDER_EMAIL="you@gmail.com"
    export SENDER_PASSWORD="xxxx xxxx xxxx xxxx"   # Gmail App Password
    export RECIPIENT_EMAIL="recipient@example.com"

    python indian_streaming_content.py
"""

import csv
import logging
import os
import smtplib
import sys
import time
from collections import Counter
from datetime import datetime, timedelta, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

OTT_API_KEY = os.environ.get("OTT_DETAILS_API_KEY", "")
OTT_BASE_URL = "https://ott-details.p.rapidapi.com"
OTT_HOST = "ott-details.p.rapidapi.com"
OUTPUT_CSV = "indian_streaming_content.csv"
LOOKBACK_DAYS = 7
CSV_FIELDS = ["title", "year", "type", "languages", "platform", "country", "date_added",
              "imdb_rating"]

# Email settings (all optional — email is skipped when any is missing).
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "")
SENDER_PASSWORD = os.environ.get("SENDER_PASSWORD", "")
RECIPIENT_EMAIL = os.environ.get("RECIPIENT_EMAIL", "")
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587

# Rate-limit / throttle settings.
MAX_RETRIES = 5             # retries per request on HTTP 429
BACKOFF_BASE = 2            # exponential base (2s, 4s, 8s …)
BACKOFF_CAP = 60            # never sleep longer than this (seconds)
PAGE_DELAY = 0.5            # seconds to sleep between consecutive API calls
MAX_PAGES_PER_REGION = 20   # stop paging after this many pages per region
DETAIL_DELAY = 0.3          # seconds to sleep between /getTitleDetails calls

# Regions and the platform names we care about (case-insensitive match).
REGIONS = ["US", "IN"]
TARGET_PLATFORMS = {
    "netflix", "prime video", "amazon prime video",
    "hulu", "hotstar", "disney+ hotstar", "jiocinema",
    "zee5", "zee 5",
}

# Readable Indian language names (OTT Details returns full names, not codes).
INDIAN_LANGUAGE_NAMES = {
    "hindi", "tamil", "telugu", "malayalam", "kannada", "bengali",
    "marathi", "gujarati", "punjabi", "odia", "assamese", "urdu",
    "sanskrit", "nepali", "sindhi", "konkani", "manipuri", "dogri",
    "santali", "maithili", "kashmiri", "bhojpuri",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# API helpers (OTT Details)
# ---------------------------------------------------------------------------

class RateLimitExhausted(Exception):
    """All retries exhausted on HTTP 429 — caller should handle gracefully."""


def _ott_headers():
    return {
        "X-RapidAPI-Key": OTT_API_KEY,
        "X-RapidAPI-Host": OTT_HOST,
    }


def call_with_backoff(method, url, *, params=None, headers=None,
                      timeout=30, max_retries=MAX_RETRIES):
    """HTTP request with Retry-After / exponential back-off on 429.

    Returns a ``requests.Response`` on success.
    Raises ``RateLimitExhausted`` if all retries fail on 429.
    Raises ``requests.HTTPError`` for other HTTP errors.
    """
    for attempt in range(1, max_retries + 1):
        resp = requests.request(
            method, url, params=params, headers=headers, timeout=timeout,
        )
        if resp.status_code != 429:
            resp.raise_for_status()
            return resp

        # --- 429 handling ---
        retry_after = resp.headers.get("Retry-After")
        if retry_after:
            try:
                wait = min(float(retry_after), BACKOFF_CAP)
            except ValueError:
                wait = min(BACKOFF_BASE ** attempt, BACKOFF_CAP)
        else:
            wait = min(BACKOFF_BASE ** attempt, BACKOFF_CAP)

        log.warning(
            "Rate-limited (429) on %s — retry %d/%d in %.1fs",
            url, attempt, max_retries, wait,
        )
        time.sleep(wait)

    # All retries exhausted
    log.error(
        "Rate limit persists after %d retries for %s — giving up.",
        max_retries, url,
    )
    raise RateLimitExhausted(f"429 after {max_retries} retries: {url}")


# ---------------------------------------------------------------------------
# OTT Details: /getnew — paginated new arrivals per region
# ---------------------------------------------------------------------------

def fetch_new_arrivals(region):
    """Page through /getnew for *region* and return a list of raw title dicts.

    Pagination is capped at ``MAX_PAGES_PER_REGION``.  A delay of
    ``PAGE_DELAY`` seconds is inserted between pages.  If rate limiting
    exhausts retries mid-pagination, already-collected results are returned.
    """
    all_results = []
    page = 1

    while page <= MAX_PAGES_PER_REGION:
        try:
            resp = call_with_backoff(
                "GET", f"{OTT_BASE_URL}/getnew",
                params={"region": region, "page": str(page)},
                headers=_ott_headers(),
            )
        except RateLimitExhausted:
            log.warning(
                "Rate limit exhausted for %s after %d page(s) — "
                "continuing with %d title(s) already collected.",
                region, page - 1, len(all_results),
            )
            break

        data = resp.json()
        results = data.get("results", [])

        if len(results) <= 1:
            # API signals end-of-data with an empty or single-element page.
            all_results.extend(results)
            break

        all_results.extend(results)
        log.info("  %s page %d → %d title(s)", region, page, len(results))
        page += 1
        time.sleep(PAGE_DELAY)

    return all_results


# ---------------------------------------------------------------------------
# OTT Details: /getTitleDetails — enrich with IMDb rating & genres
# ---------------------------------------------------------------------------

def fetch_title_details(imdb_ids):
    """Fetch /getTitleDetails for each unique IMDb id.

    Returns ``{imdb_id: detail_dict}`` with in-memory caching.
    """
    cache = {}
    ids = sorted(set(imdb_ids))
    if not ids:
        return cache

    log.info("Enriching %d unique title(s) via /getTitleDetails …", len(ids))
    for i, imdb_id in enumerate(ids):
        try:
            resp = call_with_backoff(
                "GET", f"{OTT_BASE_URL}/getTitleDetails",
                params={"imdbid": imdb_id},
                headers=_ott_headers(),
            )
            cache[imdb_id] = resp.json()
        except RateLimitExhausted:
            log.warning(
                "Rate limit hit during enrichment after %d/%d ids — "
                "remaining titles will have no rating.",
                i, len(ids),
            )
            break
        except requests.HTTPError as exc:
            log.warning("getTitleDetails failed for %s: %s", imdb_id, exc)

        if i < len(ids) - 1:
            time.sleep(DETAIL_DELAY)

    return cache


def _parse_imdb_rating(detail):
    """Extract IMDb rating as float from a title-detail dict, or None."""
    for key in ("imdbrating", "imdbRating"):
        raw = detail.get(key, "")
        if raw and raw != "N/A":
            try:
                return float(raw)
            except (ValueError, TypeError):
                pass
    return None


def _parse_genres(detail):
    """Return a set of lower-case genre strings from a title-detail dict."""
    genres = detail.get("genre") or detail.get("genres") or []
    if isinstance(genres, str):
        genres = [g.strip() for g in genres.split(",")]
    return {g.strip().lower() for g in genres if g}


# ---------------------------------------------------------------------------
# Filtering & extraction helpers
# ---------------------------------------------------------------------------

def _normalise_platform(name):
    """Return a canonical platform name for display."""
    low = name.strip().lower()
    mapping = {
        "netflix": "Netflix",
        "amazon prime video": "Prime Video",
        "prime video": "Prime Video",
        "hulu": "Hulu",
        "hotstar": "Hotstar",
        "disney+ hotstar": "Hotstar",
        "jiocinema": "Hotstar",
        "zee5": "Zee5",
        "zee 5": "Zee5",
    }
    return mapping.get(low, name.strip())


def _is_target_platform(name):
    """Return True if *name* matches one of our target streaming services."""
    return name.strip().lower() in TARGET_PLATFORMS


def is_indian_content(languages):
    """Return True if any language in *languages* (list of str) is Indian."""
    return any(lang.strip().lower() in INDIAN_LANGUAGE_NAMES for lang in languages)


def classify_type(raw_type, genres=None):
    """Map OTT Details type + genres to the user-facing content type."""
    if genres and "documentary" in genres:
        return "documentary"
    if raw_type and raw_type.strip().lower() == "series":
        return "web-series"
    return "movie"


def readable_languages(languages):
    """Return a sorted, comma-separated, title-cased language string."""
    seen = []
    for lang in languages:
        name = lang.strip().title()
        if name and name not in seen:
            seen.append(name)
    return ", ".join(sorted(seen))


def _fmt_rating(val):
    """Format a rating for HTML display: one decimal or '–'."""
    if val is None:
        return "\u2013"  # en-dash
    return f"{val:.1f}"


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(rows, path):
    """Write *rows* (list of dicts) to an .xlsx file with basic formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "New Releases"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Header row
    for col_idx, field in enumerate(CSV_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(CSV_FIELDS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))

    # Auto-width (approximate)
    for col_idx, field in enumerate(CSV_FIELDS, start=1):
        max_len = len(field)
        for row in rows:
            val = str(row.get(field, ""))
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    wb.save(path)
    log.info("Wrote Excel file to %s", path)


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def _build_summary(rows):
    """Return a string summarising release counts by country and platform."""
    by_country = Counter()
    by_combo = Counter()
    for r in rows:
        by_country[r["country"]] += 1
        by_combo[(r["country"], r["platform"])] += 1

    lines = [f"Total new Indian releases found: {len(rows)}"]
    for country in ("US", "IN"):
        if by_country[country]:
            lines.append(f"\n  {country}: {by_country[country]} title(s)")
            for (c, p), n in sorted(by_combo.items()):
                if c == country:
                    lines.append(f"    - {p}: {n}")
    return "\n".join(lines)


def _is_tamil(row):
    """Return True if the row's languages field includes Tamil."""
    langs_lower = row.get("languages", "").lower()
    tokens = {t.strip() for t in langs_lower.split(",")}
    return "tamil" in tokens or "ta" in tokens


def _build_html_body(rows):
    """Build an HTML email body with summary, top-5 US, and Tamil releases."""
    today = datetime.now(timezone.utc).strftime("%B %d, %Y")

    # --- summary counts ---
    by_country = Counter()
    by_combo = Counter()
    for r in rows:
        by_country[r["country"]] += 1
        by_combo[(r["country"], r["platform"])] += 1

    summary_rows = ""
    for country in ("US", "IN"):
        for (c, p), n in sorted(by_combo.items()):
            if c == country:
                summary_rows += (
                    f"<tr><td>{c}</td><td>{p}</td>"
                    f"<td style='text-align:center'>{n}</td></tr>\n"
                )

    # --- top 5 US releases ---
    us_rows = [r for r in rows if r["country"] == "US"][:5]
    top5_html = ""
    for r in us_rows:
        top5_html += (
            f"<tr>"
            f"<td>{r['title']}</td>"
            f"<td style='text-align:center'>{r['year']}</td>"
            f"<td>{r['type']}</td>"
            f"<td>{r['platform']}</td>"
            f"<td style='text-align:center'>{r['date_added']}</td>"
            f"<td style='text-align:center'>{_fmt_rating(r.get('imdb_rating'))}</td>"
            f"</tr>\n"
        )
    if not top5_html:
        top5_html = "<tr><td colspan='6' style='text-align:center'>No US releases found</td></tr>"

    # --- top Tamil releases: US first → IN → others, newest first ---
    tamil_country_order = {"US": 0, "IN": 1}
    tamil_rows = [r for r in rows if _is_tamil(r)]
    tamil_rows.sort(key=lambda r: r["date_added"], reverse=True)
    tamil_rows.sort(key=lambda r: tamil_country_order.get(r["country"], 99))
    if tamil_rows:
        tamil_html = ""
        for r in tamil_rows:
            tamil_html += (
                f"<tr>"
                f"<td>{r['title']}</td>"
                f"<td style='text-align:center'>{r['year']}</td>"
                f"<td>{r['type']}</td>"
                f"<td>{r['platform']}</td>"
                f"<td>{r['country']}</td>"
                f"<td style='text-align:center'>{r['date_added']}</td>"
                f"<td>{r['languages']}</td>"
                f"<td style='text-align:center'>{_fmt_rating(r.get('imdb_rating'))}</td>"
                f"</tr>\n"
            )
        tamil_section = f"""\
<h3>Top Tamil Releases</h3>
<table border="1" cellpadding="6" cellspacing="0"
       style="border-collapse:collapse;min-width:680px">
  <tr style="background:#4472C4;color:#fff">
    <th>Title</th><th>Year</th><th>Type</th><th>Platform</th>
    <th>Country</th><th>Date Added</th><th>Languages</th><th>IMDb</th>
  </tr>
  {tamil_html}
</table>"""
    else:
        tamil_section = """\
<h3>Top Tamil Releases</h3>
<p><em>No new Tamil releases in this period.</em></p>"""

    return f"""\
<html>
<body style="font-family:Arial,sans-serif;color:#333">
<h2>New Indian Movies &amp; Shows &mdash; {today}</h2>

<h3>Summary by Platform &amp; Country</h3>
<table border="1" cellpadding="6" cellspacing="0"
       style="border-collapse:collapse;min-width:320px">
  <tr style="background:#4472C4;color:#fff">
    <th>Country</th><th>Platform</th><th>Count</th>
  </tr>
  {summary_rows}
  <tr style="font-weight:bold">
    <td colspan="2">Grand Total</td>
    <td style="text-align:center">{len(rows)}</td>
  </tr>
</table>

<h3>Top 5 US Releases</h3>
<table border="1" cellpadding="6" cellspacing="0"
       style="border-collapse:collapse;min-width:520px">
  <tr style="background:#4472C4;color:#fff">
    <th>Title</th><th>Year</th><th>Type</th><th>Platform</th><th>Date Added</th><th>IMDb</th>
  </tr>
  {top5_html}
</table>

{tamil_section}

<p style="margin-top:18px;font-size:0.9em;color:#888">
  Full data is attached as an Excel file.
</p>
</body>
</html>"""


def send_email(rows, excel_path):
    """Send the results email with the Excel attachment via Gmail SMTP/TLS."""
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    msg = MIMEMultipart("mixed")
    msg["From"] = SENDER_EMAIL
    msg["To"] = RECIPIENT_EMAIL
    msg["Subject"] = f"New Indian Movies & Shows - {today_str}"

    # HTML body
    html_body = _build_html_body(rows)
    msg.attach(MIMEText(html_body, "html"))

    # Excel attachment
    attachment_name = f"new_releases_{today_str}.xlsx"
    with open(excel_path, "rb") as fh:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={attachment_name}")
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, RECIPIENT_EMAIL, msg.as_string())
        log.info("Email sent successfully to %s", RECIPIENT_EMAIL)
    except smtplib.SMTPAuthenticationError:
        log.error(
            "SMTP authentication failed. Verify SENDER_EMAIL and "
            "SENDER_PASSWORD (must be a Gmail App Password, not your "
            "regular password)."
        )
    except smtplib.SMTPException as exc:
        log.error("Failed to send email: %s", exc)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not OTT_API_KEY:
        log.error(
            "OTT_DETAILS_API_KEY environment variable is not set. "
            "Export it before running:\n  export OTT_DETAILS_API_KEY='your-key'"
        )
        sys.exit(1)

    log.info("Using OTT Details API (ott-details.p.rapidapi.com)")

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    cutoff = datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)
    log.info(
        "Fetching new arrivals for regions: %s  (last %d days)",
        ", ".join(REGIONS), LOOKBACK_DAYS,
    )

    # ------------------------------------------------------------------
    # Phase 1: Collect raw titles from /getnew per region
    # ------------------------------------------------------------------
    raw_titles = []  # list of (region, title_dict)
    for idx, region in enumerate(REGIONS):
        log.info("Querying /getnew for %s …", region)
        try:
            results = fetch_new_arrivals(region)
        except requests.HTTPError as exc:
            log.error("API error for %s: %s — skipping region.", region, exc)
            continue

        log.info("  %s: %d raw title(s) returned", region, len(results))
        for title in results:
            raw_titles.append((region, title))

        # Throttle between regions
        if idx < len(REGIONS) - 1:
            time.sleep(PAGE_DELAY)

    # ------------------------------------------------------------------
    # Phase 2: Flatten to (imdb_id, platform, country) rows, filtering
    #          for target platforms and Indian content
    # ------------------------------------------------------------------
    rows = []
    seen = set()
    imdb_ids_for_enrichment = set()

    for region, title in raw_titles:
        imdb_id = title.get("imdbid", "")
        if not imdb_id:
            continue

        languages = title.get("language") or []
        if isinstance(languages, str):
            languages = [languages]

        if not is_indian_content(languages):
            continue

        # Extract platforms for this region from streamingAvailability
        streaming = title.get("streamingAvailability") or {}
        country_map = streaming.get("country") or {}
        platforms_for_region = country_map.get(region, [])

        # If the API nests under a different key casing, try lower-case too
        if not platforms_for_region:
            platforms_for_region = country_map.get(region.lower(), [])
        # Fallback: use the first available country key
        if not platforms_for_region and country_map:
            first_key = next(iter(country_map))
            platforms_for_region = country_map[first_key]

        for entry in platforms_for_region:
            plat_name = entry.get("platform", "")
            if not _is_target_platform(plat_name):
                continue

            canonical = _normalise_platform(plat_name)
            key = (imdb_id, canonical, region)
            if key in seen:
                continue
            seen.add(key)

            imdb_ids_for_enrichment.add(imdb_id)
            rows.append({
                "title": title.get("title", ""),
                "year": title.get("released") or "",
                "type": title.get("type", "movie"),  # enriched later
                "languages": readable_languages(languages),
                "platform": canonical,
                "country": region.upper(),
                "date_added": today_str,
                "imdb_id": imdb_id,
                "imdb_rating": None,
            })

    log.info(
        "After filtering: %d row(s) across %d unique title(s)",
        len(rows), len(imdb_ids_for_enrichment),
    )

    # ------------------------------------------------------------------
    # Phase 3: Enrich with /getTitleDetails (IMDb rating + genres)
    # ------------------------------------------------------------------
    details = fetch_title_details(imdb_ids_for_enrichment)

    for r in rows:
        detail = details.get(r["imdb_id"])
        if not detail:
            continue
        r["imdb_rating"] = _parse_imdb_rating(detail)
        genres = _parse_genres(detail)
        r["type"] = classify_type(r["type"], genres)

    # ------------------------------------------------------------------
    # Phase 4: Sort, write CSV / Excel, send email
    # ------------------------------------------------------------------
    country_order = {"US": 0, "IN": 1}
    platform_order = {"Netflix": 0, "Prime Video": 1, "Hulu": 2,
                      "Hotstar": 3, "Zee5": 4}

    rows.sort(key=lambda r: r["date_added"], reverse=True)
    rows.sort(key=lambda r: (
        country_order.get(r["country"], 99),
        platform_order.get(r["platform"], 99),
    ))

    # Write CSV
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            csv_row = {k: (v if v is not None else "") for k, v in r.items()}
            writer.writerow(csv_row)

    log.info("Wrote %d record(s) to %s", len(rows), OUTPUT_CSV)

    # Write Excel and (optionally) email it
    excel_path = f"new_releases_{today_str}.xlsx"
    write_excel(rows, excel_path)

    if SENDER_EMAIL and SENDER_PASSWORD and RECIPIENT_EMAIL:
        send_email(rows, excel_path)
    else:
        log.info(
            "Email skipped — set SENDER_EMAIL, SENDER_PASSWORD, and "
            "RECIPIENT_EMAIL to enable."
        )


if __name__ == "__main__":
    main()
